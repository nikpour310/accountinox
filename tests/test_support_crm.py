from io import BytesIO

import pytest
from django.contrib import admin
from django.contrib.auth.models import Group, User
from django.core.management import call_command
from django.test import Client, RequestFactory
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.support.models import ChatMessage, ChatSession, SupportContact, SupportRating


def _set_support_session_state(client, contact, support_session):
    session = client.session
    session['support_contact_id'] = contact.id
    session['support_session_id'] = support_session.id
    session.save()


@pytest.mark.django_db
def test_support_contact_admin_export_xlsx_action():
    admin_user = User.objects.create_superuser(
        username='admin_export',
        email='admin_export@example.com',
        password='pass123456',
    )
    contact1 = SupportContact.objects.create(name='Contact One', phone='09120001001')
    contact2 = SupportContact.objects.create(name='Contact Two', phone='09120001002')

    session1 = ChatSession.objects.create(contact=contact1, user_name=contact1.name, user_phone=contact1.phone)
    ChatMessage.objects.create(session=session1, name=contact1.name, message='hello', is_from_user=True)
    ChatMessage.objects.create(session=session1, name='op', message='reply', is_from_user=False)

    session2 = ChatSession.objects.create(
        contact=contact2,
        user_name=contact2.name,
        user_phone=contact2.phone,
        is_active=False,
        closed_at=timezone.now(),
    )
    ChatMessage.objects.create(session=session2, name=contact2.name, message='help', is_from_user=True)

    client = Client()
    client.force_login(admin_user)
    response = client.post(
        reverse('admin:support_supportcontact_changelist'),
        {
            'action': 'export_contacts_xlsx',
            '_selected_action': [str(contact1.id), str(contact2.id)],
            'index': '0',
            'select_across': '0',
        },
    )

    assert response.status_code == 200
    assert (
        response['Content-Type']
        == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    worksheet = workbook.active
    headers = [worksheet.cell(row=1, column=i).value for i in range(1, 11)]
    assert headers == [
        'id',
        'name',
        'phone',
        'created_at',
        'last_seen',
        'total_sessions',
        'total_messages_user',
        'total_messages_operator',
        'last_session_status',
        'last_message_at',
    ]
    assert worksheet.max_row >= 3


@pytest.mark.django_db
def test_rate_closed_session_success():
    agent = User.objects.create_user(
        username='agent_rate_success',
        password='pass123',
        is_staff=True,
    )
    contact = SupportContact.objects.create(name='Rate User', phone='09123334444')
    support_session = ChatSession.objects.create(
        contact=contact,
        user_name=contact.name,
        user_phone=contact.phone,
        is_active=False,
        closed_at=timezone.now(),
        assigned_to=agent,
        operator=agent,
        closed_by=agent,
    )

    client = Client()
    _set_support_session_state(client, contact, support_session)
    response = client.post(
        reverse('support:rate_session', args=[support_session.id]),
        {'score': '5', 'reason': ''},
    )

    assert response.status_code == 302
    rating = SupportRating.objects.get(session=support_session)
    assert rating.score == 5
    assert rating.agent == agent


@pytest.mark.django_db
def test_rate_score_one_requires_reason():
    agent = User.objects.create_user(
        username='agent_rate_one',
        password='pass123',
        is_staff=True,
    )
    contact = SupportContact.objects.create(name='Rate User One', phone='09124445555')
    support_session = ChatSession.objects.create(
        contact=contact,
        user_name=contact.name,
        user_phone=contact.phone,
        is_active=False,
        closed_at=timezone.now(),
        assigned_to=agent,
        operator=agent,
        closed_by=agent,
    )

    client = Client()
    _set_support_session_state(client, contact, support_session)
    response = client.post(
        reverse('support:rate_session', args=[support_session.id]),
        {'score': '1', 'reason': ''},
    )

    assert response.status_code == 400
    assert SupportRating.objects.filter(session=support_session).count() == 0


@pytest.mark.django_db
def test_rate_open_session_not_allowed():
    agent = User.objects.create_user(
        username='agent_rate_open',
        password='pass123',
        is_staff=True,
    )
    contact = SupportContact.objects.create(name='Open Session User', phone='09125556666')
    support_session = ChatSession.objects.create(
        contact=contact,
        user_name=contact.name,
        user_phone=contact.phone,
        is_active=True,
        assigned_to=agent,
        operator=agent,
    )

    client = Client()
    _set_support_session_state(client, contact, support_session)
    response = client.post(
        reverse('support:rate_session', args=[support_session.id]),
        {'score': '5', 'reason': ''},
    )

    assert response.status_code == 400
    assert SupportRating.objects.filter(session=support_session).count() == 0


@pytest.mark.django_db
def test_duplicate_rating_returns_conflict():
    agent = User.objects.create_user(
        username='agent_rate_duplicate',
        password='pass123',
        is_staff=True,
    )
    contact = SupportContact.objects.create(name='Dup Rate User', phone='09126667777')
    support_session = ChatSession.objects.create(
        contact=contact,
        user_name=contact.name,
        user_phone=contact.phone,
        is_active=False,
        closed_at=timezone.now(),
        assigned_to=agent,
        operator=agent,
        closed_by=agent,
    )
    SupportRating.objects.create(session=support_session, agent=agent, score=4, reason='good')

    client = Client()
    _set_support_session_state(client, contact, support_session)
    response = client.post(
        reverse('support:rate_session', args=[support_session.id]),
        {'score': '5', 'reason': ''},
    )

    assert response.status_code == 409
    assert SupportRating.objects.filter(session=support_session).count() == 1


@pytest.mark.django_db
def test_support_rating_admin_queryset_owner_vs_agent():
    agent1 = User.objects.create_user(username='agent1', password='pass123', is_staff=True)
    agent2 = User.objects.create_user(username='agent2', password='pass123', is_staff=True)
    owner = User.objects.create_user(username='owner_user', password='pass123', is_staff=True)
    owner_group, _ = Group.objects.get_or_create(name='Owner')
    owner.groups.add(owner_group)

    contact1 = SupportContact.objects.create(name='A One', phone='09127778881')
    contact2 = SupportContact.objects.create(name='A Two', phone='09127778882')
    session1 = ChatSession.objects.create(
        contact=contact1,
        user_name=contact1.name,
        user_phone=contact1.phone,
        is_active=False,
        closed_at=timezone.now(),
        assigned_to=agent1,
        operator=agent1,
        closed_by=agent1,
    )
    session2 = ChatSession.objects.create(
        contact=contact2,
        user_name=contact2.name,
        user_phone=contact2.phone,
        is_active=False,
        closed_at=timezone.now(),
        assigned_to=agent2,
        operator=agent2,
        closed_by=agent2,
    )
    rating1 = SupportRating.objects.create(session=session1, agent=agent1, score=5, reason='')
    rating2 = SupportRating.objects.create(session=session2, agent=agent2, score=2, reason='delay')

    model_admin = admin.site._registry[SupportRating]
    request_factory = RequestFactory()

    agent_request = request_factory.get('/admin/support/supportrating/')
    agent_request.user = agent1
    agent_qs_ids = set(model_admin.get_queryset(agent_request).values_list('id', flat=True))
    assert agent_qs_ids == {rating1.id}

    owner_request = request_factory.get('/admin/support/supportrating/')
    owner_request.user = owner
    owner_qs_ids = set(model_admin.get_queryset(owner_request).values_list('id', flat=True))
    assert owner_qs_ids == {rating1.id, rating2.id}


@pytest.mark.django_db
def test_setup_support_roles_command_creates_expected_groups():
    call_command('setup_support_roles')
    group_names = set(Group.objects.values_list('name', flat=True))
    assert {'Content Admin', 'Support Agent', 'CRM Admin', 'Owner'}.issubset(group_names)
    crm_group = Group.objects.get(name='CRM Admin')
    assert crm_group.permissions.filter(codename='can_export_support_contacts').exists()


@pytest.mark.django_db
def test_message_after_close_creates_new_session_and_reappears_in_operator_unread():
    operator = User.objects.create_user(username='operator_reopen', password='pass123', is_staff=True)
    contact = SupportContact.objects.create(name='Reopen User', phone='09129990000')
    closed_session = ChatSession.objects.create(
        contact=contact,
        user_name=contact.name,
        user_phone=contact.phone,
        is_active=False,
        closed_at=timezone.now(),
    )

    user_client = Client()
    _set_support_session_state(user_client, contact, closed_session)
    response = user_client.post(
        reverse('support:send_message'),
        {'session_id': closed_session.id, 'message': 'new message after close'},
    )
    assert response.status_code == 200
    data = response.json()
    assert int(data['session_id']) != closed_session.id
    new_session_id = int(data['session_id'])

    operator_client = Client()
    operator_client.force_login(operator)
    dashboard = operator_client.get(reverse('support:operator_dashboard'))
    assert dashboard.status_code == 200
    active_ids = {session.id for session in dashboard.context['active_sessions']}
    assert new_session_id in active_ids
    assert dashboard.context['unread_count'] >= 1
